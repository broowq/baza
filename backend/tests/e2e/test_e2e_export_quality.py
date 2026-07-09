"""Экспорт — витрина качества для ЛПР клиента.

Аудит 09.07: CSV уходил клиенту с maps://-плейсхолдерами в колонке «Сайт»
(65% лидов), служебным «relevance=NN; demo=true» в заметках и без UTF-8 BOM
(кириллица в Excel — кракозябрами). Регресс закрывает формат обоих экспортов.
"""
from __future__ import annotations

import csv
import io

from app.api.routes.leads import _export_notes, _export_website
from app.models import Lead


class _L:
    """Утиный лид для юнитов хелперов."""
    def __init__(self, website="", notes=""):
        self.website = website
        self.notes = notes


def test_export_website_rewrites_2gis_placeholder():
    assert _export_website(_L(website="maps://2gis/422740746085414")) == \
        "https://2gis.ru/firm/422740746085414"
    assert _export_website(_L(website="maps://yandex/abc")) == ""
    assert _export_website(_L(website="https://real-site.ru")) == "https://real-site.ru"
    assert _export_website(_L(website="")) == ""


def test_export_notes_strips_machine_prefixes():
    assert _export_notes(_L(notes="relevance=57; demo=true; Реальный сниппет")) == "Реальный сниппет"
    assert _export_notes(_L(notes="relevance=80; Продают кедр оптом")) == "Продают кедр оптом"
    assert _export_notes(_L(notes="Обычная заметка")) == "Обычная заметка"
    assert _export_notes(_L(notes="")) == ""


def test_csv_export_has_bom_clean_sites_and_new_columns(paid_account, stub_sources, new_project):
    acct = paid_account
    project = new_project(acct)
    pid = project["id"]
    collect = acct.post(f"/api/leads/project/{pid}/collect", json={"lead_limit": 5})
    assert collect.status_code in (200, 201), collect.text

    export = acct.get(f"/api/leads/project/{pid}/export")
    assert export.status_code == 200, export.text
    raw = export.content.decode("utf-8")

    # BOM — Excel открывает кириллицу корректно.
    assert raw.startswith("﻿"), "CSV обязан начинаться с UTF-8 BOM"

    rows = list(csv.reader(io.StringIO(raw.lstrip("﻿"))))
    header, data = rows[0], rows[1:]
    assert "email_status" in header, header
    assert "description" in header, header
    assert data, "экспорт содержит лиды"

    site_i = header.index("website")
    for r in data:
        assert not r[site_i].startswith("maps://"), \
            f"служебный плейсхолдер утёк в выгрузку: {r[site_i]}"


def test_xlsx_export_has_new_columns_and_clean_notes(paid_account, stub_sources, new_project):
    from openpyxl import load_workbook

    acct = paid_account
    project = new_project(acct)
    pid = project["id"]
    collect = acct.post(f"/api/leads/project/{pid}/collect", json={"lead_limit": 5})
    assert collect.status_code in (200, 201), collect.text

    export = acct.get(f"/api/leads/project/{pid}/export.xlsx")
    assert export.status_code == 200, export.text
    wb = load_workbook(io.BytesIO(export.content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert "Email статус" in header, header
    assert "О компании" in header, header
    notes_col = header.index("Заметка") + 1
    site_col = header.index("Сайт") + 1
    for row in ws.iter_rows(min_row=2):
        note = row[notes_col - 1].value or ""
        assert "relevance=" not in str(note), f"машинный префикс в заметке: {note!r}"
        site = row[site_col - 1].value or ""
        assert not str(site).startswith("maps://"), f"плейсхолдер в сайте: {site!r}"
