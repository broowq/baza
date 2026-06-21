"""Pure (DB-free) helpers for bulk lead import (CSV / XLSX).

Responsibilities:
  * parse_upload(filename, content) -> (headers, rows): decode bytes into a
    list of original headers + a list of row dicts keyed by ORIGINAL header.
  * auto_map(headers) -> (mapping, unmapped): match headers to lead fields via
    a synonym table.
  * build_lead_dicts(rows, mapping) -> list[dict]: project each row onto the
    canonical lead fields (company/city/website/email/phone/address/notes).

No database access here on purpose — the route owns the DB, dedup and inserts.
"""

from __future__ import annotations

import csv
import io
import re

# Canonical lead field -> list of lowercased header synonyms it can match.
HEADER_SYNONYMS: dict[str, list[str]] = {
    "company": ["компания", "название", "организация", "company", "name", "firm", "фирма"],
    "city": ["город", "city", "town"],
    "website": ["сайт", "website", "site", "url", "вебсайт", "домен", "domain"],
    "email": ["email", "почта", "e-mail", "mail", "емейл", "электроннаяпочта"],
    "phone": ["телефон", "phone", "тел", "mobile", "моб", "номер"],
    "address": ["адрес", "address", "location"],
    "notes": ["заметки", "примечание", "примечания", "notes", "comment", "комментарий"],
}

# Canonical fields exposed in a built lead dict, in a stable order.
LEAD_FIELDS = ("company", "city", "website", "email", "phone", "address", "notes")


def _norm_header(h: str) -> str:
    """Normalise a header for matching: lowercase, drop spaces/punctuation.

    "E-mail адрес" -> "emailадрес"; "  Телефон " -> "телефон". This lets a
    messy human header line up with a synonym regardless of spacing/punctuation.
    """
    h = (h or "").strip().lower().replace("ё", "е")
    return re.sub(r"[\s\W_]+", "", h, flags=re.UNICODE)


# Pre-normalise synonyms once so matching is a cheap set lookup.
_NORM_SYNONYMS: dict[str, set[str]] = {
    field: {_norm_header(s) for s in syns} for field, syns in HEADER_SYNONYMS.items()
}


def _decode_csv_bytes(content: bytes) -> str:
    """Decode CSV bytes, stripping a UTF-8 BOM. Russian Excel CSVs are often
    cp1251, so fall back to that when utf-8 fails."""
    if content.startswith(b"\xef\xbb\xbf"):
        content = content[3:]
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    # Last resort — never raise here; replace undecodable bytes.
    return content.decode("utf-8", errors="replace")


def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
    text = _decode_csv_bytes(content).lstrip("﻿")
    if not text.strip():
        return [], []
    # Sniff delimiter (Russian Excel often emits ";"); fall back to ",".
    sample = text[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        if ";" in sample and sample.count(";") >= sample.count(","):
            delimiter = ";"
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        first = next(reader)
    except StopIteration:
        return [], []
    headers = [(h or "").strip() for h in first]
    rows: list[dict] = []
    for raw in reader:
        if not any((cell or "").strip() for cell in raw):
            continue  # skip fully-blank lines
        row: dict[str, str] = {}
        for i, header in enumerate(headers):
            row[header] = (raw[i].strip() if i < len(raw) and raw[i] is not None else "")
        rows.append(row)
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [("" if c is None else str(c)).strip() for c in header_row]
    rows: list[dict] = []
    for raw in rows_iter:
        if raw is None:
            continue
        cells = ["" if c is None else str(c).strip() for c in raw]
        if not any(cells):
            continue  # skip fully-blank rows
        row: dict[str, str] = {}
        for i, header in enumerate(headers):
            row[header] = cells[i] if i < len(cells) else ""
        rows.append(row)
    try:
        wb.close()
    except Exception:
        pass
    return headers, rows


def parse_upload(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    """Decode an uploaded CSV/XLSX into (headers, rows).

    rows is a list of dicts keyed by the ORIGINAL header text. Cell values are
    stringified and stripped; missing cells become "". Dispatch is by filename
    extension (.xlsx -> openpyxl, else CSV).
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def auto_map(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    """Map canonical fields to the original headers that matched them.

    Returns (mapping, unmapped) where mapping is field -> original_header and
    unmapped is the list of original headers that matched no field. First match
    wins per field (so a duplicate synonym column is ignored); a header that
    maps to an already-filled field is treated as unmapped.
    """
    mapping: dict[str, str] = {}
    unmapped: list[str] = []
    for header in headers:
        norm = _norm_header(header)
        matched_field = None
        if norm:
            for field, syns in _NORM_SYNONYMS.items():
                if field in mapping:
                    continue
                if norm in syns:
                    matched_field = field
                    break
        if matched_field is not None:
            mapping[matched_field] = header
        else:
            unmapped.append(header)
    return mapping, unmapped


def build_lead_dicts(rows: list[dict], mapping: dict[str, str]) -> list[dict]:
    """Project each raw row onto canonical lead fields.

    Each output dict has all of LEAD_FIELDS as keys (stripped strings; a field
    with no mapped column or a missing cell becomes "").
    """
    out: list[dict] = []
    for row in rows:
        lead: dict[str, str] = {}
        for field in LEAD_FIELDS:
            header = mapping.get(field)
            value = row.get(header, "") if header else ""
            lead[field] = (value or "").strip() if isinstance(value, str) else str(value or "").strip()
        out.append(lead)
    return out
