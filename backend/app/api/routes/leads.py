import csv
import io
import json
import re
import uuid
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_org, get_current_user, require_org_roles
from app.db.session import get_db
from app.models import (
    CollectionJob,
    JobStatus,
    Lead,
    LeadCallNote,
    LeadStatus,
    Membership,
    Organization,
    OutreachMessage,
    Project,
    User,
)
from app.schemas.crm import BulkLeadAction, BulkResult
from app.schemas.leads import (
    CallNoteCreate,
    CallNoteOut,
    CollectionJobOut,
    EnrichSelectedRequest,
    LeadCreateIn,
    LeadDetailOut,
    LeadEmailIn,
    LeadImportResult,
    LeadImportRowError,
    LeadOut,
    LeadTouchIn,
    LeadUpdate,
    LeadWarehouseRef,
    PaginatedLeadsOut,
    RunCollectionRequest,
)
from app.services import company_warehouse, lead_import
from app.services.crm import log_activity, stage_label
from app.services.quota import ensure_lead_quota
from app.services.scoring import score_lead
from app.services.audit import log_action
from app.tasks.jobs import collect_leads_task, enrich_leads_task
from app.utils.url_tools import extract_domain, is_aggregator_domain, is_real_domain


def _clip(value: str, max_len: int) -> str:
    return (value or "")[:max_len]


def _build_website_and_domain(raw: str) -> tuple[str, str]:
    """Resolve a user-supplied website into (website, domain).

    Empty input → a unique stable placeholder ("manual://<uuid>") so the
    NOT NULL + unique(project_id, website) constraint is satisfied without a
    real URL. A bare domain ("acme.ru") gets an https:// scheme. domain is only
    extracted for a real http(s) site whose host is a real, non-aggregator
    domain; otherwise "".
    """
    raw = (raw or "").strip()
    if not raw:
        return f"manual://{uuid.uuid4().hex}", ""
    website = raw
    if "://" not in website:
        website = f"https://{website}"
    domain = ""
    if website.startswith(("http://", "https://")):
        d = extract_domain(website)
        if d and is_real_domain(d) and not is_aggregator_domain(d):
            domain = d
    return _clip(website, 300), _clip(domain, 255)


def _find_duplicate_lead_id(db: Session, project_id, *, company: str, city: str, domain: str, website: str):
    """Return an existing lead id if this lead duplicates one in the project.

    Mirrors the collector's rule: with a real domain → dup if website OR domain
    matches; without a domain → dup if (company AND city) match case-insensitively.
    """
    if domain:
        dup = db.execute(
            select(Lead.id).where(
                Lead.project_id == project_id,
                (Lead.website == website) | (Lead.domain == domain),
            )
        ).first()
    else:
        dup = db.execute(
            select(Lead.id).where(
                Lead.project_id == project_id,
                func.lower(Lead.company) == (company or "").lower(),
                func.lower(Lead.city) == (city or "").lower(),
            )
        ).first()
    return dup[0] if dup else None


def _clean_tags(tags: list[str] | None) -> list[str]:
    """Trim, dedupe (case-insensitive), cap 30 chars each, max 20 tags."""
    cleaned: list[str] = []
    seen: set[str] = set()
    for t in tags or []:
        t = (t or "").strip()[:30]
        if t and t.lower() not in seen:
            seen.add(t.lower())
            cleaned.append(t)
        if len(cleaned) >= 20:
            break
    return cleaned


def _ci_contains(col, term: str):
    """Collation-independent case-insensitive substring match for lead search.

    Plain ILIKE / lower() only case-fold ASCII under a C/POSIX database
    collation, so a capitalised Cyrillic company ("Юникорн") could never be
    found by a lowercase query — broken for a Russian-first CRM. Case-fold both
    sides via the ICU root collation (`und-x-icu`), which folds Unicode case
    regardless of the database's LC_CTYPE.
    """
    return func.lower(col.collate("und-x-icu")).like(f"%{term.lower()}%")


def _fire_lead_webhook_best_effort(db: Session, organization: Organization, lead: Lead, project: Project) -> None:
    """Reuse the existing CRM webhook (Bitrix24/AmoCRM/custom) for a new lead.

    Fire-and-forget; never blocks or fails the request. Mirrors the payload the
    enrichment loop pushes in app/tasks/jobs.py.
    """
    try:
        if not organization.lead_webhook_url:
            return
        from app.tasks.webhook_tasks import push_lead_webhook

        payload = {
            "id": str(lead.id),
            "company": lead.company,
            "city": lead.city,
            "email": lead.email,
            "phone": lead.phone,
            "address": lead.address,
            "website": lead.website,
            "score": lead.score,
            "status": lead.status.value,
            "tags": lead.tags or [],
            "project_id": str(lead.project_id),
            "project_name": project.name if project else "",
        }
        push_lead_webhook.delay(organization.lead_webhook_url, payload)
    except Exception:
        pass

router = APIRouter(prefix="/leads", tags=["leads"])

MAX_CONCURRENT_JOBS_PER_ORG = 3


def _count_active_org_jobs(db: Session, organization_id) -> int:
    """Count running/queued jobs across all projects in an organization."""
    return db.scalar(
        select(func.count(CollectionJob.id)).where(
            CollectionJob.organization_id == organization_id,
            CollectionJob.status.in_([JobStatus.queued, JobStatus.running]),
        )
    ) or 0


def _lock_project_and_check_active_job(db: Session, project_id, kind: str) -> CollectionJob | None:
    """Acquire a row-level lock on the Project row, then check for an active job of the given kind.

    Using SELECT … FOR UPDATE serialises concurrent requests against the same project:
    the second transaction blocks until the first commits (inserting the new job), then
    re-reads — finding the job — and returns it.  Without the lock, two simultaneous
    POSTs could both pass the check before either committed and spawn duplicate jobs.
    """
    # Lock the project row for the duration of this transaction.
    db.execute(
        select(Project).where(Project.id == project_id).with_for_update()
    )
    # Now safely check for an existing active job — no concurrent transaction can
    # insert a competing job for this project while we hold the lock.
    return db.execute(
        select(CollectionJob).where(
            CollectionJob.project_id == project_id,
            CollectionJob.kind == kind,
            CollectionJob.status.in_([JobStatus.queued, JobStatus.running]),
        )
    ).scalar_one_or_none()


@router.get("/project/{project_id}", response_model=list[LeadOut])
def list_project_leads(
    project_id: str,
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")
    return db.execute(select(Lead).where(Lead.project_id == project.id).order_by(Lead.created_at.desc()).limit(5000)).scalars().all()


@router.get("/project/{project_id}/table", response_model=PaginatedLeadsOut)
def list_project_leads_table(
    project_id: str,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=25, ge=1, le=200),
    q: str = "",
    status: LeadStatus | None = None,
    has_email: bool | None = None,
    has_phone: bool | None = None,
    min_score: int | None = Query(default=None, ge=0, le=100),
    max_score: int | None = Query(default=None, ge=0, le=100),
    assigned_to: str | None = Query(
        default=None,
        description='User UUID, "me" (current user), or "none"/"unassigned" (no owner)',
    ),
    sort: str = "score",
    order: str = "desc",
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    query = select(Lead).where(Lead.project_id == project.id)
    count_query = select(func.count(Lead.id)).where(Lead.project_id == project.id)

    if q.strip():
        term = q.strip()
        search_cond = (
            _ci_contains(Lead.company, term)
            | _ci_contains(Lead.website, term)
            | _ci_contains(Lead.city, term)
            | _ci_contains(Lead.domain, term)
            | _ci_contains(Lead.email, term)
            | _ci_contains(Lead.phone, term)
            | _ci_contains(Lead.address, term)
        )
        query = query.where(search_cond)
        count_query = count_query.where(search_cond)
    if status:
        query = query.where(Lead.status == status)
        count_query = count_query.where(Lead.status == status)
    if has_email is not None:
        if has_email:
            query = query.where(Lead.email != "")
            count_query = count_query.where(Lead.email != "")
        else:
            query = query.where(Lead.email == "")
            count_query = count_query.where(Lead.email == "")
    if has_phone is not None:
        if has_phone:
            query = query.where(Lead.phone != "")
            count_query = count_query.where(Lead.phone != "")
        else:
            query = query.where(Lead.phone == "")
            count_query = count_query.where(Lead.phone == "")
    if min_score is not None:
        query = query.where(Lead.score >= min_score)
        count_query = count_query.where(Lead.score >= min_score)
    if max_score is not None:
        query = query.where(Lead.score <= max_score)
        count_query = count_query.where(Lead.score <= max_score)
    if assigned_to is not None and assigned_to != "":
        token = assigned_to.strip().lower()
        if token in ("none", "unassigned"):
            query = query.where(Lead.assigned_to_user_id.is_(None))
            count_query = count_query.where(Lead.assigned_to_user_id.is_(None))
        elif token == "me":
            query = query.where(Lead.assigned_to_user_id == user.id)
            count_query = count_query.where(Lead.assigned_to_user_id == user.id)
        else:
            # An explicit user UUID. Malformed → 422 (matches the API's strict
            # param validation elsewhere); an unknown-but-valid UUID simply
            # matches nothing (no cross-org leak — leads stay project-scoped).
            try:
                assignee_id = uuid.UUID(assigned_to)
            except ValueError as exc:
                raise HTTPException(
                    status_code=422, detail="Некорректный идентификатор пользователя"
                ) from exc
            query = query.where(Lead.assigned_to_user_id == assignee_id)
            count_query = count_query.where(Lead.assigned_to_user_id == assignee_id)

    sort_column = {
        "score": Lead.score,
        "company": Lead.company,
        "status": Lead.status,
        "created_at": Lead.created_at,
    }.get(sort, Lead.created_at)
    # Deterministic ordering with tiebreakers. Without these, ORDER BY score
    # (with many tied scores) returned a DIFFERENT order on every refetch — the
    # 6-sec poll made cards visibly reshuffle and a fresh batch of 10 looked
    # "перемешанным" with the old ones. created_at DESC groups the newest leads
    # at the top of each tie band; id is the final stable tiebreaker.
    order_cols = [sort_column.asc() if order == "asc" else sort_column.desc()]
    if sort != "created_at":
        order_cols.append(Lead.created_at.desc())
    order_cols.append(Lead.id)
    query = query.order_by(*order_cols)

    total = db.scalar(count_query) or 0
    items = db.execute(query.offset((page - 1) * per_page).limit(per_page)).scalars().all()
    return PaginatedLeadsOut(items=items, total=total, page=page, per_page=per_page)


@router.get("/all", response_model=PaginatedLeadsOut)
def list_org_leads_table(
    search: str = "",
    status: str = "all",
    project_id: UUID | None = None,
    assigned_to: str = "all",  # "all" | "unassigned" | <user UUID>
    sort: str = "score",  # score | created_at | last_contacted_at | company
    order: str = "desc",
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=1, le=100),
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    """Org-wide leads list across ALL (non-deleted) projects.

    Mirrors list_project_leads_table's filter/sort/pagination logic, but scopes
    by organization (joined to Project, excluding soft-deleted projects) instead
    of a single project, and stamps each row with its project_name.
    """
    # Base scope: org's leads joined to a live (non-deleted) project. The join
    # itself enforces "project belongs to org AND is not soft-deleted".
    base = (
        select(Lead)
        .join(Project, Project.id == Lead.project_id)
        .where(
            Lead.organization_id == organization.id,
            Project.organization_id == organization.id,
            Project.deleted_at.is_(None),
        )
    )
    count_base = (
        select(func.count(Lead.id))
        .join(Project, Project.id == Lead.project_id)
        .where(
            Lead.organization_id == organization.id,
            Project.organization_id == organization.id,
            Project.deleted_at.is_(None),
        )
    )

    # Optional single-project filter — must belong to this org (and be live).
    if project_id is not None:
        proj = db.get(Project, project_id)
        if (
            not proj
            or proj.organization_id != organization.id
            or proj.deleted_at is not None
        ):
            raise HTTPException(status_code=404, detail="Проект не найден")
        base = base.where(Lead.project_id == project_id)
        count_base = count_base.where(Lead.project_id == project_id)

    # Case-insensitive search across the main contact columns. Require >=2 chars
    # so a stray single keystroke doesn't trigger a full ILIKE scan.
    term = search.strip()
    if len(term) >= 2:
        search_clause = or_(
            _ci_contains(Lead.company, term),
            _ci_contains(Lead.email, term),
            _ci_contains(Lead.phone, term),
            _ci_contains(Lead.website, term),
            _ci_contains(Lead.domain, term),
            _ci_contains(Lead.city, term),
        )
        base = base.where(search_clause)
        count_base = count_base.where(search_clause)

    if status != "all":
        base = base.where(Lead.status == status)
        count_base = count_base.where(Lead.status == status)

    token = (assigned_to or "all").strip().lower()
    if token == "unassigned":
        base = base.where(Lead.assigned_to_user_id.is_(None))
        count_base = count_base.where(Lead.assigned_to_user_id.is_(None))
    elif token != "all":
        # Explicit user UUID. Malformed → 422; an unknown-but-valid UUID simply
        # matches nothing (no cross-org leak — leads stay org-scoped).
        try:
            assignee_id = uuid.UUID(assigned_to)
        except ValueError as exc:
            raise HTTPException(
                status_code=422, detail="Некорректный идентификатор пользователя"
            ) from exc
        base = base.where(Lead.assigned_to_user_id == assignee_id)
        count_base = count_base.where(Lead.assigned_to_user_id == assignee_id)

    sort_column = {
        "score": Lead.score,
        "created_at": Lead.created_at,
        "last_contacted_at": Lead.last_contacted_at,
        "company": Lead.company,
    }.get(sort, Lead.score)
    order_cols = [sort_column.asc() if order == "asc" else sort_column.desc()]
    # Stable secondary ordering so tied rows don't reshuffle between refetches.
    if sort != "created_at":
        order_cols.append(Lead.created_at.desc())
    order_cols.append(Lead.id)
    base = base.order_by(*order_cols)

    total = db.scalar(count_base) or 0
    leads = (
        db.execute(base.offset((page - 1) * per_page).limit(per_page))
        .scalars()
        .all()
    )

    # Resolve project_name for just the page's projects — one extra query, no
    # N+1 (a {project_id: name} map keyed by the leads actually on this page).
    project_ids = {lead.project_id for lead in leads}
    name_by_project: dict[UUID, str] = {}
    if project_ids:
        rows = db.execute(
            select(Project.id, Project.name).where(Project.id.in_(project_ids))
        ).all()
        name_by_project = {pid: name for pid, name in rows}

    items = [
        LeadOut.model_validate(lead).model_copy(
            update={"project_name": name_by_project.get(lead.project_id, "")}
        )
        for lead in leads
    ]
    return PaginatedLeadsOut(items=items, total=total, page=page, per_page=per_page)


@router.post("/project/{project_id}/collect", response_model=CollectionJobOut)
def run_collection(
    project_id: str,
    payload: RunCollectionRequest,
    organization: Organization = Depends(get_current_org),
    membership=Depends(require_org_roles("owner", "admin")),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")
    if _lock_project_and_check_active_job(db, project.id, "collect"):
        raise HTTPException(status_code=409, detail="Сбор уже запущен для этого проекта")
    if _count_active_org_jobs(db, organization.id) >= MAX_CONCURRENT_JOBS_PER_ORG:
        raise HTTPException(
            status_code=429,
            detail=f"Превышен лимит одновременных задач ({MAX_CONCURRENT_JOBS_PER_ORG}). Дождитесь завершения текущих задач.",
        )

    # Collect dose ceiling — the task clamps to the same 200; match it here so the
    # quota pre-check reflects what will actually be added (not the raw request).
    dose = min(payload.lead_limit, 200)
    ensure_lead_quota(organization, dose)
    job = CollectionJob(
        organization_id=organization.id,
        project_id=project.id,
        status=JobStatus.queued,
        kind="collect",
        requested_limit=dose,
    )
    db.add(job)
    db.flush()
    log_action(
        db,
        user_id=str(membership.user_id),
        organization_id=str(organization.id),
        action="leads.collect.queued",
        meta={"project_id": str(project.id), "limit": dose, "job_id": str(job.id)},
    )
    db.commit()
    db.refresh(job)
    collect_leads_task.delay(str(job.id))
    return job


@router.post("/project/{project_id}/enrich", response_model=CollectionJobOut)
def run_enrichment(
    project_id: str,
    payload: RunCollectionRequest,
    organization: Organization = Depends(get_current_org),
    membership=Depends(require_org_roles("owner", "admin")),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")
    if _lock_project_and_check_active_job(db, project.id, "enrich"):
        raise HTTPException(status_code=409, detail="Обогащение уже запущено для этого проекта")
    if _count_active_org_jobs(db, organization.id) >= MAX_CONCURRENT_JOBS_PER_ORG:
        raise HTTPException(
            status_code=429,
            detail=f"Превышен лимит одновременных задач ({MAX_CONCURRENT_JOBS_PER_ORG}). Дождитесь завершения текущих задач.",
        )
    # A lead is enrichable if it was never enriched, OR it still has no
    # actionable contact (no email AND no phone). We deliberately do NOT require
    # the address to also be empty: warehouse/2GIS leads often carry an address
    # but no email/phone, and gating on address left them permanently
    # un-enrichable ("Нет лидов для обогащения") even though they need contacts.
    enrichable = db.scalar(
        select(func.count(Lead.id)).where(
            Lead.project_id == project.id,
            or_(
                Lead.enriched.is_(False),
                (Lead.email == "") & (Lead.phone == ""),
            ),
        )
    ) or 0
    if enrichable == 0:
        raise HTTPException(status_code=400, detail="Нет лидов для обогащения")
    job = CollectionJob(
        organization_id=organization.id,
        project_id=project.id,
        status=JobStatus.queued,
        kind="enrich",
        requested_limit=payload.lead_limit,
    )
    db.add(job)
    db.flush()
    log_action(
        db,
        user_id=str(membership.user_id),
        organization_id=str(organization.id),
        action="leads.enrich.queued",
        meta={"project_id": str(project.id), "limit": payload.lead_limit, "job_id": str(job.id)},
    )
    db.commit()
    db.refresh(job)
    enrich_leads_task.delay(str(job.id))
    return job


@router.post("/project/{project_id}/enrich-selected", response_model=CollectionJobOut)
def run_selected_enrichment(
    project_id: str,
    payload: EnrichSelectedRequest,
    organization: Organization = Depends(get_current_org),
    membership=Depends(require_org_roles("owner", "admin")),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")
    if _lock_project_and_check_active_job(db, project.id, "enrich"):
        raise HTTPException(status_code=409, detail="Обогащение уже запущено для этого проекта")
    if _count_active_org_jobs(db, organization.id) >= MAX_CONCURRENT_JOBS_PER_ORG:
        raise HTTPException(
            status_code=429,
            detail=f"Превышен лимит одновременных задач ({MAX_CONCURRENT_JOBS_PER_ORG}). Дождитесь завершения текущих задач.",
        )
    if not payload.lead_ids:
        raise HTTPException(status_code=400, detail="Не выбраны лиды для обогащения")
    safe_ids: list[UUID] = []
    for raw_id in payload.lead_ids:
        try:
            safe_ids.append(UUID(raw_id))
        except ValueError:
            continue
    if not safe_ids:
        raise HTTPException(status_code=400, detail="Переданы некорректные идентификаторы лидов")
    valid_count = db.scalar(select(func.count(Lead.id)).where(Lead.project_id == project.id, Lead.id.in_(safe_ids))) or 0
    if valid_count == 0:
        raise HTTPException(status_code=400, detail="Выбранные лиды не найдены в проекте")
    job = CollectionJob(
        organization_id=organization.id,
        project_id=project.id,
        status=JobStatus.queued,
        kind="enrich",
        requested_limit=valid_count,
    )
    db.add(job)
    db.flush()
    log_action(
        db,
        user_id=str(membership.user_id),
        organization_id=str(organization.id),
        action="leads.enrich_selected.queued",
        meta={"project_id": str(project.id), "lead_ids_count": len(safe_ids), "job_id": str(job.id)},
    )
    db.commit()
    db.refresh(job)
    enrich_leads_task.delay(str(job.id), [str(item) for item in safe_ids])
    return job


@router.get("/jobs/project/{project_id}", response_model=list[CollectionJobOut])
def list_jobs(
    project_id: str,
    limit: int = Query(default=50, ge=1, le=500),
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    """List recent collection/enrichment jobs (default last 50, max 500)."""
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")
    return (
        db.execute(
            select(CollectionJob)
            .where(CollectionJob.project_id == project.id)
            .order_by(CollectionJob.created_at.desc())
            .limit(limit)
        )
        .scalars()
        .all()
    )


# ── Хелперы экспорта: выгрузка — витрина качества для ЛПР клиента ──────────
# Аудит 09.07: в колонку «Сайт» уходил служебный плейсхолдер maps://2gis/…
# (65% лидов), в заметки — префикс «relevance=57; demo=true», а CSV без BOM
# открывался в Excel кракозябрами.

_NOTES_MACHINE_PREFIX_RE = re.compile(r"^(?:relevance=\d+;\s*)?(?:demo=true;\s*)?")


def _export_website(lead: Lead) -> str:
    """Человеческий сайт для выгрузки: maps://2gis/{id} → ссылка на карточку
    2ГИС, прочие maps:// (нет публичного URL) → пусто."""
    w = (lead.website or "").strip()
    if w.startswith("maps://2gis/"):
        firm_id = w[len("maps://2gis/"):]
        return f"https://2gis.ru/firm/{firm_id}" if firm_id.isdigit() else ""
    if w.startswith("maps://"):
        return ""
    return w


def _export_notes(lead: Lead) -> str:
    """Заметка без машинных префиксов сборщика."""
    return _NOTES_MACHINE_PREFIX_RE.sub("", lead.notes or "").strip()


@router.get("/project/{project_id}/export")
def export_project_csv(
    project_id: str,
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    safe_name = "".join(ch if ch.isascii() and (ch.isalnum() or ch in "-_") else "-" for ch in project.name.lower())
    safe_name = "-".join(part for part in safe_name.split("-") if part) or "project"
    file_name = f"leads-{safe_name}-{datetime.now(timezone.utc).date().isoformat()}.csv"

    def _generate_csv():
        output = io.StringIO()
        # UTF-8 BOM: без него Excel (главный инструмент РОПа) открывает
        # кириллицу кракозябрами. Излишен для парсеров — они BOM игнорируют.
        output.write("\ufeff")
        writer = csv.writer(output)
        writer.writerow(
            [
                "company",
                "city",
                "website",
                "domain",
                "email",
                "email_status",
                "phone",
                "address",
                "description",
                "score",
                "status",
                "source_url",
                "contacts_json",
                "demo",
            ]
        )
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        for lead in db.execute(
            select(Lead).where(Lead.project_id == project.id)
        ).yield_per(500).scalars():
            output.seek(0)
            output.truncate(0)
            contacts = lead.contacts_json or lead.contacts or {}
            writer.writerow(
                [
                    lead.company,
                    lead.city,
                    _export_website(lead),
                    lead.domain or extract_domain(lead.website),
                    lead.email,
                    lead.email_status,
                    lead.phone,
                    lead.address,
                    (lead.description or "")[:600],
                    lead.score,
                    lead.status.value,
                    lead.source_url,
                    json.dumps(contacts, ensure_ascii=False),
                    str(bool(lead.demo)).lower(),
                ]
            )
            yield output.getvalue()

    return StreamingResponse(
        _generate_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@router.get("/project/{project_id}/export.xlsx")
def export_project_xlsx(
    project_id: str,
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    """Excel export (.xlsx) — opens correctly in Excel (unlike CSV with UTF-8 BOM).

    Uses openpyxl. Columns are Russian, phones/emails are hyperlinks where possible.
    For projects > 10k leads, CSV endpoint is still preferable (memory-friendly stream).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    safe_name = "".join(ch if ch.isascii() and (ch.isalnum() or ch in "-_") else "-" for ch in project.name.lower())
    safe_name = "-".join(part for part in safe_name.split("-") if part) or "project"
    file_name = f"leads-{safe_name}-{datetime.now(timezone.utc).date().isoformat()}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    headers = [
        ("Компания", 40), ("Город", 18), ("Сайт", 30), ("Email", 28),
        ("Email статус", 13), ("Телефон", 18), ("Адрес", 40),
        ("О компании", 45), ("Score", 8), ("Статус", 14),
        ("Теги", 20), ("Последний контакт", 16), ("Напомнить", 16),
        ("Заметка", 40),
    ]
    for i, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F6FBD")
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    status_labels = {
        "new": "Новый", "contacted": "Связались",
        "qualified": "Квалифицирован", "proposal": "КП отправлено",
        "won": "Сделка", "rejected": "Отклонён",
    }
    email_status_labels = {
        "valid": "валиден", "no_mx": "домен мёртв",
        "syntax": "опечатка", "skipped": "не проверен", "": "",
    }

    row_num = 2
    for lead in db.execute(select(Lead).where(Lead.project_id == project.id)).yield_per(500).scalars():
        ws.cell(row=row_num, column=1, value=lead.company or "")
        ws.cell(row=row_num, column=2, value=lead.city or "")
        export_site = _export_website(lead)
        website_cell = ws.cell(row=row_num, column=3, value=export_site)
        if export_site.startswith("http"):
            website_cell.hyperlink = export_site
            website_cell.font = Font(color="0000FF", underline="single")
        email_cell = ws.cell(row=row_num, column=4, value=lead.email or "")
        if lead.email:
            email_cell.hyperlink = f"mailto:{lead.email}"
            email_cell.font = Font(color="0000FF", underline="single")
        ws.cell(row=row_num, column=5, value=email_status_labels.get(lead.email_status, lead.email_status))
        phone_cell = ws.cell(row=row_num, column=6, value=lead.phone or "")
        if lead.phone:
            phone_cell.hyperlink = f"tel:{lead.phone}"
            phone_cell.font = Font(color="0000FF", underline="single")
        ws.cell(row=row_num, column=7, value=lead.address or "")
        ws.cell(row=row_num, column=8, value=(lead.description or "")[:600])
        ws.cell(row=row_num, column=9, value=lead.score)
        ws.cell(row=row_num, column=10, value=status_labels.get(lead.status.value, lead.status.value))
        ws.cell(row=row_num, column=11, value=", ".join(lead.tags or []))
        if lead.last_contacted_at:
            ws.cell(row=row_num, column=12, value=lead.last_contacted_at.strftime("%d.%m.%Y"))
        if lead.reminder_at:
            ws.cell(row=row_num, column=13, value=lead.reminder_at.strftime("%d.%m.%Y"))
        ws.cell(row=row_num, column=14, value=_export_notes(lead))
        row_num += 1

    # Write to in-memory buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


# Full pipeline: new|contacted|qualified|proposal|won|rejected (derived from the
# LeadStatus enum so a new stage automatically becomes accepted everywhere).
VALID_STATUSES = {s.value for s in LeadStatus}

# Hard cap on rows accepted per import — guards against memory/DB blow-up from a
# malicious or accidental giant upload.
MAX_IMPORT_ROWS = 5000


@router.post("/project/{project_id}", response_model=LeadOut, status_code=201)
def create_lead(
    project_id: str,
    payload: LeadCreateIn,
    organization: Organization = Depends(get_current_org),
    membership=Depends(require_org_roles("owner", "admin")),
    db: Session = Depends(get_db),
):
    """Manually add a single lead to a project (canManage gate).

    The lead is the user's OWN data, so it does NOT consume the AI-collection
    quota (organization.leads_used_current_month is left untouched).
    """
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    if payload.status not in VALID_STATUSES:
        raise HTTPException(
            status_code=422,
            detail=f"Недопустимый статус. Допустимые значения: {', '.join(sorted(VALID_STATUSES))}",
        )

    # An assignee, if given, must be a member of THIS org (else 422).
    if payload.assigned_to_user_id is not None:
        is_member = db.scalar(
            select(func.count(Membership.id)).where(
                Membership.organization_id == organization.id,
                Membership.user_id == payload.assigned_to_user_id,
            )
        ) or 0
        if not is_member:
            raise HTTPException(status_code=422, detail="Пользователь не в организации")

    company = _clip(payload.company.strip(), 180)
    city = _clip(payload.city.strip(), 120)
    website, domain = _build_website_and_domain(payload.website)

    dup_id = _find_duplicate_lead_id(
        db, project.id, company=company, city=city, domain=domain, website=website
    )
    if dup_id is not None:
        raise HTTPException(
            status_code=409,
            detail={
                "detail": "Такой лид уже есть в проекте",
                "existing_lead_id": str(dup_id),
            },
        )

    score = score_lead(
        domain=domain,
        company=company,
        niche=project.niche,
        has_email=bool(payload.email.strip()),
        has_phone=bool(payload.phone.strip()),
        has_address=bool(payload.address.strip()),
        demo=False,
    )

    lead = Lead(
        organization_id=organization.id,
        project_id=project.id,
        company=company,
        city=city,
        website=website,
        domain=domain,
        email=_clip(payload.email.strip(), 255),
        phone=_clip(payload.phone.strip(), 80),
        address=_clip(payload.address.strip(), 300),
        notes=payload.notes or "",
        tags=_clean_tags(payload.tags),
        status=LeadStatus(payload.status),
        deal_value=payload.deal_value,
        assigned_to_user_id=payload.assigned_to_user_id,
        score=score,
        source="manual",
        enriched=False,
    )
    db.add(lead)
    db.flush()
    log_activity(db, lead=lead, kind="created", text="Лид добавлен вручную")
    db.commit()
    db.refresh(lead)

    _fire_lead_webhook_best_effort(db, organization, lead, project)
    return lead


@router.post("/project/{project_id}/import", response_model=LeadImportResult)
def import_leads(
    project_id: str,
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    organization: Organization = Depends(get_current_org),
    membership=Depends(require_org_roles("owner", "admin")),
    db: Session = Depends(get_db),
):
    """Bulk-import leads from a CSV/XLSX upload (canManage gate).

    dry_run=True previews counts + column mapping + a small sample of leads that
    WOULD be created, inserting nothing. Imported leads are the user's own data
    and do NOT consume the AI-collection quota.
    """
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    filename = file.filename or ""
    lower = filename.lower()
    if not (lower.endswith(".csv") or lower.endswith(".xlsx")):
        raise HTTPException(status_code=422, detail="Поддерживаются только файлы .csv или .xlsx")

    content = file.file.read()
    if not content:
        raise HTTPException(status_code=422, detail="Пустой файл")

    try:
        headers, rows = lead_import.parse_upload(filename, content)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Не удалось прочитать файл") from exc

    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=422, detail=f"Слишком большой файл (макс {MAX_IMPORT_ROWS} строк)")

    mapping, unmapped = lead_import.auto_map(headers)
    lead_dicts = lead_import.build_lead_dicts(rows, mapping)

    errors: list[LeadImportRowError] = []
    duplicates = 0
    created = 0
    sample: list[Lead] = []
    # In-batch dedup keys so two identical rows in the same file don't both insert.
    seen_domain: set[str] = set()
    seen_website: set[str] = set()
    seen_company_city: set[tuple[str, str]] = set()

    for idx, raw in enumerate(lead_dicts):
        # Header is row 1; first data row is spreadsheet row 2.
        row_num = idx + 2
        company = _clip((raw.get("company") or "").strip(), 180)
        if not company:
            errors.append(LeadImportRowError(row=row_num, error="Пустая компания"))
            continue
        city = _clip((raw.get("city") or "").strip(), 120)
        website, domain = _build_website_and_domain(raw.get("website") or "")

        # Dedup within the batch first, then against existing project leads.
        if domain:
            if domain in seen_domain or website in seen_website:
                duplicates += 1
                continue
        else:
            key = (company.lower(), city.lower())
            if key in seen_company_city:
                duplicates += 1
                continue
        if _find_duplicate_lead_id(
            db, project.id, company=company, city=city, domain=domain, website=website
        ) is not None:
            duplicates += 1
            continue

        # Record batch keys now that this row is accepted as new.
        if domain:
            seen_domain.add(domain)
            seen_website.add(website)
        else:
            seen_company_city.add((company.lower(), city.lower()))

        email = _clip((raw.get("email") or "").strip(), 255)
        phone = _clip((raw.get("phone") or "").strip(), 80)
        address = _clip((raw.get("address") or "").strip(), 300)
        notes = (raw.get("notes") or "")[:10000]
        score = score_lead(
            domain=domain,
            company=company,
            niche=project.niche,
            has_email=bool(email),
            has_phone=bool(phone),
            has_address=bool(address),
            demo=False,
        )
        lead = Lead(
            organization_id=organization.id,
            project_id=project.id,
            company=company,
            city=city,
            website=website,
            domain=domain,
            email=email,
            phone=phone,
            address=address,
            notes=notes,
            tags=[],
            status=LeadStatus.new,
            score=score,
            source="manual",
            enriched=False,
        )
        created += 1
        if not dry_run:
            db.add(lead)
        if len(sample) < 5:
            sample.append(lead)

    if not dry_run and created:
        db.commit()

    # Build LeadOut samples. For dry_run the leads are unsaved (no id/created_at),
    # so serialize a LeadOut-shaped dict rather than model_validate(lead).
    sample_out: list[LeadOut] = []
    if dry_run:
        for lead in sample:
            sample_out.append(
                LeadOut(
                    id=uuid.uuid4(),
                    organization_id=organization.id,
                    project_id=project.id,
                    company=lead.company,
                    city=lead.city,
                    website=lead.website,
                    domain=lead.domain,
                    email=lead.email,
                    email_status="",
                    phone=lead.phone,
                    address=lead.address,
                    contacts={},
                    contacts_json={},
                    score=lead.score,
                    notes=lead.notes or "",
                    tags=[],
                    status=lead.status,
                    deal_value=0,
                    source_url="",
                    source=lead.source,
                    external_id="",
                    enriched=False,
                    demo=False,
                    created_at=datetime.now(timezone.utc),
                )
            )
    else:
        for lead in sample:
            db.refresh(lead)
            sample_out.append(LeadOut.model_validate(lead))

    return LeadImportResult(
        total=len(lead_dicts),
        created=created,
        duplicates=duplicates,
        errors=errors,
        dry_run=dry_run,
        detected_columns=mapping,
        unmapped_headers=unmapped,
        sample=sample_out,
    )


# Russian labels for the data source, used in the composed description.
_SOURCE_LABELS_RU = {
    "2gis": "2ГИС",
    "yandex_maps": "Яндекс.Карты",
    "rusprofile": "Rusprofile",
    "searxng": "веб-поиск",
    "yandex_search": "Яндекс.Поиск",
    "bing": "Bing",
    "maps_searxng": "картам",
    "warehouse": "базе компаний",
}


def _compose_lead_description(lead: Lead, company) -> str:
    """Build a human-readable description for a lead that has none of its own.

    Composes from categories (warehouse) + city + source + contact availability.
    Returns "" only when there is genuinely nothing to say. If the lead's notes
    carry a real description (after stripping the internal "relevance=…;" /
    "demo=…;" prefixes the collector stores there), prefer that instead.
    """
    # Приоритет — РЕАЛЬНЫЙ текст о компании, а не суррогат из метаданных:
    # 1) lead.description (meta-description сайта с обогащения / описание
    #    кандидата со сбора / бэкфилл со склада);
    if (lead.description or "").strip():
        return lead.description.strip()[:600]

    # 2) сниппет, сохранённый в notes при сборе (за машинными префиксами).
    # The collector stores machine prefixes in notes ("relevance=NN; demo=true; ")
    # followed by the real snippet. Strip the prefixes to recover the snippet.
    note = _NOTES_MACHINE_PREFIX_RE.sub("", lead.notes or "").strip()
    if note:
        return note[:600]

    # 3) описание этой компании со склада (могло появиться после сбора —
    #    например, из обогащения этой же компании другой организацией).
    wh_desc = (getattr(company, "description", "") or "").strip() if company else ""
    if wh_desc:
        return wh_desc[:600]

    # 4) суррогат из категорий/метаданных — лучше, чем пусто.
    parts: list[str] = []
    categories = list(getattr(company, "categories", []) or []) if company else []
    if categories:
        parts.append(", ".join(categories[:5]))
    if lead.city:
        parts.append(f"г. {lead.city}" if not lead.city.lower().startswith("г.") else lead.city)
    if lead.source:
        label = _SOURCE_LABELS_RU.get(lead.source, lead.source)
        parts.append(f"источник: {label}")

    have = []
    if lead.phone:
        have.append("телефон")
    if lead.email:
        have.append("email")
    if lead.address:
        have.append("адрес")
    if lead.website and not lead.website.startswith("maps://"):
        have.append("сайт")
    if have:
        parts.append("есть " + ", ".join(have))
    else:
        parts.append("контакты не найдены")

    return ". ".join(p for p in parts if p).strip()


@router.get("/{lead_id}", response_model=LeadDetailOut)
def get_lead_detail(
    lead_id: uuid.UUID,
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    """Full lead detail + computed description + warehouse cross-reference.

    Org-scoped: the lead (and its project) must belong to the caller's org,
    otherwise 404 (same opacity as the other lead routes — we don't leak
    existence across orgs).
    """
    lead = db.get(Lead, lead_id)
    if not lead or lead.organization_id != organization.id:
        raise HTTPException(status_code=404, detail="Лид не найден")
    project = db.get(Project, lead.project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Лид не найден")

    # Cross-reference the shared company warehouse by the same dedup_key rule.
    # Best-effort — a warehouse miss/error just yields found=False.
    company = company_warehouse.find_company_for_lead(
        db, domain=lead.domain or "", company=lead.company or "", city=lead.city or ""
    )
    warehouse_ref = LeadWarehouseRef()
    if company is not None:
        # NOTE: Company.niches is deliberately NOT exposed here — those are
        # OTHER organizations' search niches (their go-to-market intent), a
        # cross-tenant leak. sources/categories are public data and stay.
        warehouse_ref = LeadWarehouseRef(
            found=True,
            company_id=company.id,
            times_seen=company.times_seen,
            first_seen_at=company.first_seen_at,
            last_seen_at=company.last_seen_at,
            sources=list(company.sources or []),
            categories=list(company.categories or []),
            best_score=company.best_score,
            inn=company.inn or "",
            twogis_firm_id=company.twogis_firm_id or "",
        )

    description = _compose_lead_description(lead, company)

    detail = LeadDetailOut.model_validate(lead)
    detail.description = description
    detail.warehouse = warehouse_ref
    return detail


@router.patch("/{lead_id}", response_model=LeadOut)
def update_lead(
    lead_id: uuid.UUID,
    payload: LeadUpdate,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    lead = db.get(Lead, lead_id)
    if not lead or lead.organization_id != organization.id:
        raise HTTPException(status_code=404, detail="Лид не найден")
    project = db.get(Project, lead.project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Лид не найден")

    fields_set = payload.model_fields_set

    if payload.status is not None:
        if payload.status not in VALID_STATUSES:
            raise HTTPException(
                status_code=422,
                detail=f"Недопустимый статус. Допустимые значения: {', '.join(sorted(VALID_STATUSES))}",
            )
        new_status = LeadStatus(payload.status)
        if new_status != lead.status:
            old_status = lead.status
            lead.status = new_status
            log_activity(
                db,
                lead=lead,
                kind="stage_changed",
                text=f"Стадия: {stage_label(old_status)} → {stage_label(new_status)}",
                user=user,
                meta={"from": old_status.value, "to": new_status.value},
            )

    # assigned_to_user_id — model_fields_set distinguishes an explicit null
    # (unassign) from an omitted field. A UUID must belong to a member of THIS
    # org (else 422); we resolve the display name for the activity text.
    if "assigned_to_user_id" in fields_set:
        new_assignee = payload.assigned_to_user_id
        if new_assignee is None:
            if lead.assigned_to_user_id is not None:
                lead.assigned_to_user_id = None
                log_activity(db, lead=lead, kind="unassigned", text="Снято назначение", user=user)
        else:
            is_member = db.scalar(
                select(func.count(Membership.id)).where(
                    Membership.organization_id == organization.id,
                    Membership.user_id == new_assignee,
                )
            ) or 0
            if not is_member:
                raise HTTPException(status_code=422, detail="Пользователь не в организации")
            if lead.assigned_to_user_id != new_assignee:
                lead.assigned_to_user_id = new_assignee
                assignee = db.get(User, new_assignee)
                name = (assignee.full_name or assignee.email or "") if assignee else ""
                log_activity(
                    db,
                    lead=lead,
                    kind="assigned",
                    text=f"Назначен: {name}",
                    user=user,
                    meta={"assigned_to": str(new_assignee)},
                )

    if payload.notes is not None and payload.notes != lead.notes:
        lead.notes = payload.notes
        log_activity(db, lead=lead, kind="note", text="Заметка обновлена", user=user)
    if payload.tags is not None:
        # Sanitize tags: trim, dedupe, max 30 chars each
        cleaned_tags = []
        seen = set()
        for t in payload.tags:
            t = (t or "").strip()[:30]
            if t and t.lower() not in seen:
                seen.add(t.lower())
                cleaned_tags.append(t)
        lead.tags = cleaned_tags
    if payload.deal_value is not None and payload.deal_value != lead.deal_value:
        lead.deal_value = payload.deal_value
        log_activity(
            db,
            lead=lead,
            kind="value_changed",
            text=f"Сумма сделки: {payload.deal_value} ₽",
            user=user,
            meta={"deal_value": payload.deal_value},
        )
    # expected_close_at — explicit null clears it (model_fields_set guard).
    if "expected_close_at" in fields_set:
        lead.expected_close_at = payload.expected_close_at
    if payload.last_contacted_at is not None:
        lead.last_contacted_at = payload.last_contacted_at
    # Use model_fields_set so an explicit {"reminder_at": null} CLEARS the
    # reminder (the × button sends null); an omitted field leaves it untouched.
    # The old `is not None` guard silently ignored the clear → dead button.
    if "reminder_at" in fields_set:
        lead.reminder_at = payload.reminder_at
    if payload.mark_contacted:
        # Convenience flag — sets last_contacted_at=now() and bumps status to "contacted"
        # if it was still "new". Lets sales click one button after a call.
        lead.last_contacted_at = datetime.now(timezone.utc)
        if lead.status == LeadStatus.new:
            lead.status = LeadStatus.contacted
        log_activity(db, lead=lead, kind="contacted", text="Отмечен контакт", user=user)

    db.commit()
    db.refresh(lead)
    return lead


@router.post("/project/{project_id}/bulk", response_model=BulkResult)
def bulk_lead_action(
    project_id: uuid.UUID,
    payload: BulkLeadAction,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Apply one action over many leads at once (assign / stage / add_tag / delete).

    Scoped to this org + project: ids outside the project are silently skipped
    (no cross-org/-project leak). Returns the count of affected leads. assign,
    stage and add_tag write an activity per affected lead; delete does not (the
    lead — and its timeline — is gone).
    """
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    action = payload.action

    # Validate the action payload up-front (before touching any rows).
    if action == "assign":
        if payload.assigned_to_user_id is not None:
            is_member = db.scalar(
                select(func.count(Membership.id)).where(
                    Membership.organization_id == organization.id,
                    Membership.user_id == payload.assigned_to_user_id,
                )
            ) or 0
            if not is_member:
                raise HTTPException(status_code=422, detail="Пользователь не в организации")
    elif action == "stage":
        if not payload.status or payload.status not in VALID_STATUSES:
            raise HTTPException(
                status_code=422,
                detail=f"Недопустимый статус. Допустимые значения: {', '.join(sorted(VALID_STATUSES))}",
            )
    elif action == "add_tag":
        tag = (payload.tag or "").strip()[:30]
        if not tag:
            raise HTTPException(status_code=422, detail="Пустой тег")
    elif action == "delete":
        pass
    else:
        raise HTTPException(status_code=422, detail="Неизвестное действие")

    # Fetch the in-scope leads (org + project). Anything not here is ignored.
    leads = db.execute(
        select(Lead).where(
            Lead.organization_id == organization.id,
            Lead.project_id == project.id,
            Lead.id.in_(payload.lead_ids),
        )
    ).scalars().all()
    if not leads:
        return BulkResult(updated=0)

    updated = 0

    if action == "assign":
        new_assignee = payload.assigned_to_user_id
        name = ""
        if new_assignee is not None:
            assignee = db.get(User, new_assignee)
            name = (assignee.full_name or assignee.email or "") if assignee else ""
        for lead in leads:
            if lead.assigned_to_user_id == new_assignee:
                continue
            lead.assigned_to_user_id = new_assignee
            if new_assignee is None:
                log_activity(db, lead=lead, kind="unassigned", text="Снято назначение", user=user)
            else:
                log_activity(
                    db,
                    lead=lead,
                    kind="assigned",
                    text=f"Назначен: {name}",
                    user=user,
                    meta={"assigned_to": str(new_assignee)},
                )
            updated += 1

    elif action == "stage":
        new_status = LeadStatus(payload.status)
        for lead in leads:
            if lead.status == new_status:
                continue
            old_status = lead.status
            lead.status = new_status
            log_activity(
                db,
                lead=lead,
                kind="stage_changed",
                text=f"Стадия: {stage_label(old_status)} → {stage_label(new_status)}",
                user=user,
                meta={"from": old_status.value, "to": new_status.value},
            )
            updated += 1

    elif action == "add_tag":
        tag = (payload.tag or "").strip()[:30]
        for lead in leads:
            existing = list(lead.tags or [])
            if any((e or "").lower() == tag.lower() for e in existing):
                continue
            if len(existing) >= 20:
                continue
            existing.append(tag)
            lead.tags = existing
            log_activity(
                db,
                lead=lead,
                kind="tag_added",
                text=f"Тег: {tag}",
                user=user,
                meta={"tag": tag},
            )
            updated += 1

    elif action == "delete":
        for lead in leads:
            db.delete(lead)
            updated += 1

    db.commit()
    return BulkResult(updated=updated)


def _get_org_lead_or_404(db: Session, lead_id: uuid.UUID, organization: Organization) -> Lead:
    """Org-scoped lead fetch shared by the call-journal routes (404 opacity)."""
    lead = db.get(Lead, lead_id)
    if not lead or lead.organization_id != organization.id:
        raise HTTPException(status_code=404, detail="Лид не найден")
    project = db.get(Project, lead.project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Лид не найден")
    return lead


@router.get("/{lead_id}/calls", response_model=list[CallNoteOut])
def list_call_notes(
    lead_id: uuid.UUID,
    organization: Organization = Depends(get_current_org),
    db: Session = Depends(get_db),
):
    """Call journal for a lead, newest first."""
    _get_org_lead_or_404(db, lead_id, organization)
    notes = db.execute(
        select(LeadCallNote)
        .where(
            LeadCallNote.lead_id == lead_id,
            LeadCallNote.organization_id == organization.id,
        )
        .order_by(LeadCallNote.created_at.desc())
        .limit(100)
    ).scalars().all()
    return notes


@router.post("/{lead_id}/calls", response_model=CallNoteOut, status_code=201)
def add_call_note(
    lead_id: uuid.UUID,
    payload: CallNoteCreate,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Record a call on the lead: who called (current user) + optional comment.

    Side effects mirror mark_contacted: last_contacted_at=now(), and a lead
    still in "new" moves to "contacted" — so the journal is the single button
    sales clicks after dialing.
    """
    lead = _get_org_lead_or_404(db, lead_id, organization)
    note = LeadCallNote(
        organization_id=organization.id,
        lead_id=lead.id,
        user_id=user.id,
        user_name=(user.full_name or user.email or "")[:120],
        comment=(payload.comment or "").strip(),
    )
    db.add(note)
    lead.last_contacted_at = datetime.now(timezone.utc)
    if lead.status == LeadStatus.new:
        lead.status = LeadStatus.contacted
    db.commit()
    db.refresh(note)
    return note


@router.post("/{lead_id}/email")
def send_lead_email(
    lead_id: uuid.UUID,
    payload: LeadEmailIn,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Reply/write to the lead by email through the org's SMTP.

    On success the OutreachMessage surfaces in the unified timeline (via crm) —
    so we do NOT also log a duplicate LeadActivity here. Bumps last_contacted_at
    and moves a lead still in "new" to "contacted".
    """
    from app.api.routes.outreach import _get_settings_row
    from app.services import outreach

    lead = _get_org_lead_or_404(db, lead_id, organization)
    if not lead.email:
        raise HTTPException(status_code=422, detail="У лида нет email")
    if lead.email_opt_out:
        raise HTTPException(status_code=409, detail="Лид отписался от писем")

    s = _get_settings_row(db, organization.id)
    if s is None or not s.smtp_host:
        raise HTTPException(
            status_code=409,
            detail="Почта организации не настроена (Настройки → Email)",
        )

    subject = payload.subject
    tok = outreach.new_track_token()
    html = outreach.inject_tracking(outreach._to_html(payload.body), tok)
    try:
        outreach.send_via_smtp(
            s,
            to_email=lead.email,
            subject=subject,
            html_body=html,
            text_body=payload.body,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=502, detail=f"Не удалось отправить: {exc}"[:300]
        ) from exc

    message = OutreachMessage(
        organization_id=organization.id,
        lead_id=lead.id,
        enrollment_id=None,
        to_email=lead.email,
        subject=subject[:300],
        status="sent",
        track_token=tok,
    )
    db.add(message)
    lead.last_contacted_at = datetime.now(timezone.utc)
    if lead.status == LeadStatus.new:
        lead.status = LeadStatus.contacted
    db.commit()
    db.refresh(message)
    return {"id": str(message.id), "status": "sent"}


@router.post("/{lead_id}/touch")
def add_lead_touch(
    lead_id: uuid.UUID,
    payload: LeadTouchIn,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Log a one-click channel touch (call / WhatsApp / Telegram button).

    Records a "touch" activity in the unified timeline and mirrors the
    mark-contacted side effects (last_contacted_at=now(); new → contacted).
    """
    lead = _get_org_lead_or_404(db, lead_id, organization)
    channel = payload.channel
    if channel not in ("call", "whatsapp", "telegram"):
        raise HTTPException(status_code=422, detail="Недопустимый канал")

    log_activity(
        db,
        lead=lead,
        kind="touch",
        text=payload.note,
        user=user,
        meta={"channel": channel},
    )
    lead.last_contacted_at = datetime.now(timezone.utc)
    if lead.status == LeadStatus.new:
        lead.status = LeadStatus.contacted
    db.commit()
    return {"ok": True}


@router.delete("/{lead_id}", status_code=204)
def delete_lead(
    lead_id: uuid.UUID,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    lead = db.get(Lead, lead_id)
    if not lead or lead.organization_id != organization.id:
        raise HTTPException(status_code=404, detail="Лид не найден")
    project = db.get(Project, lead.project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Лид не найден")

    db.delete(lead)
    db.commit()
    return Response(status_code=204)


@router.get("/project/{project_id}/stats")
def get_project_stats(
    project_id: uuid.UUID,
    organization: Organization = Depends(get_current_org),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if not project or project.organization_id != organization.id or project.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Проект не найден")

    total = db.execute(select(func.count(Lead.id)).where(Lead.project_id == project.id)).scalar()
    enriched = db.execute(
        select(func.count(Lead.id)).where(Lead.project_id == project.id, Lead.enriched == True)
    ).scalar()
    with_email = db.execute(
        select(func.count(Lead.id)).where(
            Lead.project_id == project.id, Lead.email.isnot(None), Lead.email != ""
        )
    ).scalar()
    avg_score = (
        db.execute(select(func.avg(Lead.score)).where(Lead.project_id == project.id)).scalar() or 0
    )
    return {
        "total": total,
        "enriched": enriched,
        "with_email": with_email,
        "avg_score": round(float(avg_score), 1),
    }
